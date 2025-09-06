import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active

data = [
    ['Name', 'Age', 'Country'],
    ['John', 30, 'USA'],
    ['Mary', 25, 'USA'],
    ['Bob', 22, 'Canada']
]

for row in data:
    sheet.append(row)

workbook.save('data.xlsx')

loaded_workbook = openpyxl.load_workbook('data.xlsx')
loaded_sheet = loaded_workbook.active

for row in loaded_sheet.iter_rows(values_only=True):
    print(row)